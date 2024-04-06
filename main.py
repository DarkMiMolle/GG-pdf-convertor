import sys
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class Operation:
    def __init__(self, lines: list[str]):
        
        self.initial_lines = lines
        self.date = lines[0][:10]
        if "/" not in self.date:
            self.ref = "ERR"
            return

        if "€" in lines[0]: # only one line
            self.custom_label = "-"
            self.ref = "-"
            sign = "-"
            if "+" in lines[0]:
                sign = "+"
            
            datas = lines[0][10:].split(sign)
            
            self.global_label, self.money = ("-".join(datas[:-1]), datas[-1])
            self.money = float(sign + self.money[:-2].replace(" ", "").replace(",", ".")) # remove \xa0€
            return            
        
        if len(lines) < 3:
            self.ref = "ERR"
            return
                
        self.global_label = lines[0][10:]
        
        self.custom_label = lines[1]

        sign = "-"
        if "+" in lines[2]:
            sign = "+"
        self.ref, self.money = lines[2].split(sign)
        self.money = float(sign + self.money[:-2].replace(" ", "").replace(",", ".")) # remove \xa0€

    def iter(self) -> list[str]:
        return [self.date, self.global_label, self.custom_label, self.ref, self.money]

def generate_data_from(filename: str) -> list[list[str]]:
    pdf = PdfReader(filename)
    
    operations = []


    for page in pdf.pages:
        page_txt = page.extract_text().splitlines()
        
        from_line = 4 if len(operations) != 0 else 9

        page_operations = [Operation(page_txt[i:i+3]) for i in range(from_line, len(page_txt)-1)]

        # print()
        # page_operations = [[page_operations[i][0][:10], page_operations[i][0][10:], *page_operations[i][1:]] for i in range(len(page_operations)) if page_operations[i][-1].startswith("Ref")]
        operations = [*operations, *[operation for operation in page_operations if operation.ref != "ERR"]]
        
    # for operation in operations:
    #     ref: str = operation[-1]
    #     if not ref.startswith("Ref"):
    #         raise ValueError("expected Ref of the operation")
        
    #     op = "+" if "+" in ref else "-"
    #     ref, val = ref.split(op)
    #     val = val[:-2] + val[-1:] # remove unexpected value: \xa0
    #     operation[-1] = ref
    #     operation.append(op + val)
    
    return operations

def get_column_letter_perso(i: int) -> str:
    if i == 0: return 'A'
    
    alphabet = [chr(ord('A') + i) for i in range(26)]
    col_name = ""
    while i != 0:
        col_name += alphabet[i%26]
        i //= 26
        
    return col_name

def adjust_cols_width(sheet: Worksheet):
    for col in sheet.columns:
        letter = col[0].column_letter
        width = 0
        for cell in col:
            width = max(width, (len(str(cell.value)) + 5) * 1.25)
        
        sheet.column_dimensions[letter].width = width

def write_excel_for(operations: list[Operation]) -> Workbook:
    xl = Workbook()
    sheet: Worksheet = xl.active

    for i in range(len(operations)):
        operation = operations[i]
        for j in range(len(operation.iter())):
            col_name = get_column_letter(j+1)
            cell: Cell = sheet[f'{col_name}{i+1}']
            cell.font = Font(size=14)
            cell.value = operation.iter()[j]
            
    for cell in sheet['E']:
        cell: Cell = cell
        cell.number_format = "00.00 €"
    
    sheet.insert_rows(0, 2)
    rows = [row for row in sheet.iter_rows(0, 5)]
    rows[0][0].value = "Date"
    rows[0][1].value = "Label GG global"
    rows[0][2].value = "Label detail"
    rows[0][3].value = "Ref"
    rows[0][4].value = "Montant"
    for cell in rows[0]:
        cell.font = Font(size=14)
    
    adjust_cols_width(sheet)
        
    return xl

def main():
    filename = sys.argv[1]
    xl = write_excel_for(generate_data_from(filename))
    xl.save(filename[:-3] + "xlsx")

if __name__ == "__main__": 
    main()